# -*- coding: utf-8 -*-
# 台帳＋AgiMaレポートから prices.json を生成（型式→素体/PSA10）
import openpyxl, glob, os, re, json, datetime
DATA=r"C:\Users\PC_User\Desktop\ClaudeCode\data\Agito-Market"
OUT=r"C:\Users\PC_User\Desktop\AgiS-app\prices.json"

def latest(pat):
    fs=[f for f in glob.glob(os.path.join(DATA,pat)) if not os.path.basename(f).startswith("~$")]
    return max(fs,key=lambda f:(re.search(r"(\d{8})",f) or ["","0"])[1]) if fs else None

def slug(t):
    s=re.sub(r"[\s/]+","-",str(t).strip().lower())
    s=re.sub(r"-+","-",s).strip("-")
    return s

def num(v):
    return v if isinstance(v,(int,float)) else None

db={}
def put(typ,name,sotai,psa10,src):
    if not typ: return
    k=slug(typ)
    if not k: return
    # 既存があってPSA10が入ってればスキップ（台帳優先で先に入れる）
    if k in db and db[k].get("psa10"): return
    db[k]={"name":name,"sotai":sotai,"psa10":psa10,"src":src}

# 台帳（優先）
lg=latest("*ledger*.xlsx")
if lg:
    wb=openpyxl.load_workbook(lg,data_only=True)
    sg=[w for w in wb.worksheets if "シングル" in w.title][0]
    for r in range(2,sg.max_row+1):
        nm=sg.cell(r,3).value; typ=sg.cell(r,4).value
        if nm and typ:
            put(typ,nm,num(sg.cell(r,23).value),num(sg.cell(r,24).value),"台帳")

# AgiMaレポート ③乖離ピック / ②TOP / ①差分
rp=latest("AgiMa相場レポート_*.xlsx")
if rp:
    wb=openpyxl.load_workbook(rp,data_only=True)
    for ws in wb.worksheets:
        t=ws.title
        for r in range(2,ws.max_row+1):
            a=ws.cell(r,1).value
            if a is None or str(a).startswith("【"): continue
            if "乖離" in t:   # 判定|名|型番|美品|PSA10
                put(ws.cell(r,3).value, ws.cell(r,2).value, num(ws.cell(r,4).value), num(ws.cell(r,5).value), "GR乖離")
            elif "TOP" in t or "取引" in t:  # 順位|銘柄|型番|件数|PSA10
                put(ws.cell(r,3).value, ws.cell(r,2).value, None, num(ws.cell(r,5).value), "GR出来高")
            elif "差分" in t:  # ID|名|型式|旧|新PSA10
                put(ws.cell(r,3).value, ws.cell(r,2).value, None, num(ws.cell(r,5).value), "GR差分")

out={"updated":datetime.date(2026,7,3).isoformat(),"count":len(db),"cards":db}
with open(OUT,"w",encoding="utf-8") as f:
    json.dump(out,f,ensure_ascii=False,indent=0)
print("prices.json:",len(db),"件 ->",OUT)
